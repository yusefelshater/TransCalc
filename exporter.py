"""
Exporter utilities for TransCalc
- export_json(state, out_path)
- export_excel(state, out_path)
- export_run(state, runs_dir="runs")

Expected state structure:
{
  "timestamp": "ISO8601" (optional, will be added if missing),
  "inputs": { ... },
  "results": {
    "quantities": {
      "volume_m3": float,
      "mix_total_ton": float,
      "bitumen_ton": float,
      "rubber_ton": float,
      "aggregates_total_ton": float,
      "aggregates_breakdown": {
        "<type_id>": {"mass_ton": float, "price_per_ton": float, "subtotal": float}
      }
    },
    "costs": {
      "aggregates_subtotal": float,
      "bitumen_subtotal": float,
      "rubber_subtotal": float,
      "materials_subtotal": float,
      "overhead_total": float,
      "grand_total": float
    }
  },
  "warnings": [str, ...],
  "metadata": {"user": str, "source": "gui|cli|script", ...}
}
"""

from __future__ import annotations

import os
import json
from datetime import datetime
from typing import Any, Dict, List, Tuple
import csv

try:
    from openpyxl import Workbook
except ImportError:  # Graceful error if dependency not installed yet
    Workbook = None  # type: ignore


# ----------------------------- Helpers -----------------------------

def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _to_cell(value: Any) -> Any:
    """Convert complex values (dict/list) to compact JSON string for Excel cell safety."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _flatten(prefix: str, data: Dict[str, Any], out: Dict[str, Any]) -> None:
    for k, v in data.items():
        key = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            _flatten(key, v, out)
        else:
            out[key] = v


def _autosize(ws) -> None:
    # Autosize columns based on max length of cell content
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, length)
            except Exception:
                pass
        ws.column_dimensions[col].width = min(max(10, max_length + 2), 60)


# ----------------------------- Exports -----------------------------

def export_json(state: Dict[str, Any], out_path: str) -> str:
    """Write state snapshot to JSON file (UTF-8, pretty-printed)."""
    snap = dict(state)
    if "timestamp" not in snap:
        snap["timestamp"] = datetime.now().isoformat(timespec="seconds")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    return out_path


def export_excel(state: Dict[str, Any], out_path: str) -> str:
    """Write state snapshot to Excel workbook with multiple sheets.
    Sheets: Inputs, Results, Aggregates, Costs, Overheads, Warnings, Meta
    """
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Please install requirements first.")

    wb = Workbook()

    # Inputs sheet
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    inputs = state.get("inputs", {})
    flat_inputs: Dict[str, Any] = {}
    if isinstance(inputs, dict):
        _flatten("", inputs, flat_inputs)
    ws_inputs.append(["Key", "Value"])
    for k in sorted(flat_inputs.keys()):
        ws_inputs.append([k, _to_cell(flat_inputs[k])])
    _autosize(ws_inputs)

    # Results sheet (quantities overview)
    ws_results = wb.create_sheet("Results")
    results = state.get("results", {}) if isinstance(state.get("results", {}), dict) else {}
    quantities = results.get("quantities", {}) if isinstance(results.get("quantities", {}), dict) else {}
    ws_results.append(["Metric", "Value"])
    for metric in [
        "volume_m3",
        "mix_total_ton",
        "bitumen_ton",
        "rubber_ton",
        "aggregates_total_ton",
    ]:
        if metric in quantities:
            ws_results.append([metric, _to_cell(quantities[metric])])
    _autosize(ws_results)

    # Aggregates breakdown sheet
    ws_aggs = wb.create_sheet("Aggregates")
    ws_aggs.append(["type_id", "mass_ton", "price_per_ton", "subtotal"])
    aggs_break = {}
    if isinstance(quantities, dict):
        aggs_break = quantities.get("aggregates_breakdown", {}) or {}
    if isinstance(aggs_break, dict):
        for type_id, row in aggs_break.items():
            if isinstance(row, dict):
                ws_aggs.append([
                    type_id,
                    _to_cell(row.get("mass_ton")),
                    _to_cell(row.get("price_per_ton")),
                    _to_cell(row.get("subtotal")),
                ])
    _autosize(ws_aggs)

    # Costs sheet
    ws_costs = wb.create_sheet("Costs")
    costs = results.get("costs", {}) if isinstance(results.get("costs", {}), dict) else {}
    ws_costs.append(["Item", "Value"])
    for item in [
        "aggregates_subtotal",
        "bitumen_subtotal",
        "rubber_subtotal",
        "materials_subtotal",
        "overhead_total",
        "grand_total",
    ]:
        if item in costs:
            ws_costs.append([item, _to_cell(costs[item])])
    _autosize(ws_costs)

    # Overheads sheet
    ws_ov = wb.create_sheet("Overheads")
    ws_ov.append(["component_id", "percent", "egp_per_ton"])
    overheads = (inputs.get("overheads") or {}) if isinstance(inputs, dict) else {}
    comps = overheads.get("components") if isinstance(overheads, dict) else None
    if isinstance(comps, list):
        for comp in comps:
            if isinstance(comp, dict):
                ws_ov.append([
                    _to_cell(comp.get("id")),
                    _to_cell(comp.get("percent")),
                    _to_cell(comp.get("egp_per_ton")),
                ])
    else:
        # fallback: flatten whole overheads if no components list
        flat_ov: Dict[str, Any] = {}
        if isinstance(overheads, dict):
            _flatten("", overheads, flat_ov)
        ws_ov.append(["---", "---", "---"])
        for k in sorted(flat_ov.keys()):
            ws_ov.append([k, _to_cell(flat_ov[k]), None])
    _autosize(ws_ov)

    # Warnings sheet
    ws_warn = wb.create_sheet("Warnings")
    ws_warn.append(["warning_text"])
    for w in state.get("warnings", []) or []:
        ws_warn.append([_to_cell(w)])
    _autosize(ws_warn)

    # Meta sheet
    ws_meta = wb.create_sheet("Meta")
    meta = state.get("metadata", {}) if isinstance(state.get("metadata", {}), dict) else {}
    if "timestamp" not in state:
        state["timestamp"] = datetime.now().isoformat(timespec="seconds")
    meta_out = {
        "timestamp": state.get("timestamp"),
        **meta,
    }
    ws_meta.append(["Key", "Value"])
    for k in sorted(meta_out.keys()):
        ws_meta.append([k, _to_cell(meta_out[k])])
    _autosize(ws_meta)

    # Save workbook
    wb.save(out_path)
    return out_path


def export_run(state: Dict[str, Any], runs_dir: str = "runs") -> Dict[str, str]:
    """Export both JSON and Excel to runs/<timestamp>_transcalc_compare.*
    Returns dict with paths: {"json": path, "xlsx": path}
    """
    _ensure_dir(runs_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(runs_dir, f"{ts}_transcalc_compare")
    json_path = base + ".json"
    xlsx_path = base + ".xlsx"

    export_json(state, json_path)
    export_excel(state, xlsx_path)

    return {"json": json_path, "xlsx": xlsx_path}


# ----------------------------- Planner exports -----------------------------
def export_planner(analysis: Dict[str, Any], runs_dir: str = "runs") -> Dict[str, str]:
    """Export planner analysis dict to runs/<ts>_planner.* (JSON and CSV).
    CSV columns: type,name,lat,lon,total_score,total_score_norm,near_road,midpoint,quarry,rubber,highway,ready_mix,bitumen,landuse_label,landuse_score,buildings_count
    Returns dict with paths: {"json": path, "csv": path}
    """
    _ensure_dir(runs_dir)
    ts = _ts()
    base = os.path.join(runs_dir, f"{ts}_planner")
    json_path = base + ".json"
    csv_path = base + ".csv"

    # Write JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)

    # Build flat rows for CSV
    rows: List[Dict[str, Any]] = []
    for kind in ("existing", "proposed"):
        for item in analysis.get(kind, []) or []:
            sc = item.get("score", {}) if isinstance(item, dict) else {}
            scores = sc.get("scores", {}) if isinstance(sc, dict) else {}
            rows.append({
                "type": kind,
                "name": item.get("name"),
                "lat": item.get("lat"),
                "lon": item.get("lon"),
                "total_score": sc.get("total_score"),
                "total_score_norm": sc.get("total_score_norm"),
                "near_road": scores.get("near_road"),
                "midpoint": scores.get("midpoint"),
                "quarry": scores.get("quarry"),
                "rubber": scores.get("rubber"),
                "highway": scores.get("highway"),
                "ready_mix": scores.get("ready_mix"),
                "bitumen": scores.get("bitumen"),
                "landuse_label": scores.get("landuse_label"),
                "landuse_score": scores.get("landuse_score"),
                "buildings_count": scores.get("buildings_count"),
            })

    # Write CSV
    fieldnames = [
        "type","name","lat","lon","total_score","total_score_norm",
        "near_road","midpoint","quarry","rubber","highway","ready_mix","bitumen",
        "landuse_label","landuse_score","buildings_count"
    ]
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    return {"json": json_path, "csv": csv_path}


# ----------------------------- CLI hook (optional) -----------------------------
if __name__ == "__main__":
    # Minimal manual test placeholder
    demo_state = {
        "inputs": {
            "project": {"length_km": 1.0, "width_m": 7.5, "thickness_m": 0.05, "density_ton_per_m3": 2.35},
            "mix": {
                "bitumen_prop_of_mix": 0.055,
                "rubber_prop_of_bitumen": 0.02,
                "aggregates_shares": {"coarse": 0.4, "medium": 0.3, "fine": 0.2},
                "aggregates_type_ids": {"coarse": "coarse_s1", "medium": "coarse_s2", "fine": "sand_fine"}
            },
            "overheads": {"mode": "percent", "components": [
                {"id": "transport", "percent": 0.10},
                {"id": "waste", "percent": 0.04},
                {"id": "salaries", "percent": 0.12},
                {"id": "equip", "percent": 0.08},
                {"id": "profit", "percent": 0.06}
            ]}
        },
        "results": {
            "quantities": {
                "volume_m3": 1000.0,
                "mix_total_ton": 2350.0,
                "bitumen_ton": 129.25,
                "rubber_ton": 2.585,
                "aggregates_total_ton": 2220.75,
                "aggregates_breakdown": {
                    "coarse_s1": {"mass_ton": 940.0, "price_per_ton": 189, "subtotal": 177660.0},
                    "coarse_s2": {"mass_ton": 705.0, "price_per_ton": 167, "subtotal": 117735.0},
                    "sand_fine": {"mass_ton": 470.0, "price_per_ton": 80,  "subtotal": 37600.0}
                }
            },
            "costs": {
                "aggregates_subtotal": 333000.0,
                "bitumen_subtotal": 0.0,
                "rubber_subtotal": 0.0,
                "materials_subtotal": 0.0,
                "overhead_total": 0.0,
                "grand_total": 0.0
            }
        },
        "warnings": ["Demo warning only"],
        "metadata": {"user": "local", "source": "script"}
    }
    # Clean placeholder numeric that may confuse formatters
    demo_state["results"]["costs"]["aggregates_subtotal"] = 333000.0
    paths = export_run(demo_state)
    print("Exported:", paths)
